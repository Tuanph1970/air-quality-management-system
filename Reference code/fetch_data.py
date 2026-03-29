#!/usr/bin/env python3
"""
Fetch average air quality data from tw-envisoft.tedp.vn
- Layer 1: HTTP Basic Auth (tw-admin / tw-admin)
- Layer 2: Web form login (duongngocbach)
- Fetches all KK stations for a given province and date range
- Saves each station to a separate Excel file (or one combined file)

Usage:
    python fetch_data.py --from-date 2026-03-03 --to-date 2026-03-04
    python fetch_data.py --from-date 2026-03-01 --to-date 2026-03-31 --output-dir ./output
    python fetch_data.py --from-date 2026-03-03 --to-date 2026-03-04 --combined
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.sync_api import sync_playwright, BrowserContext, Page

# ─── Configuration ─────────────────────────────────────────────────────────────
MAIN_URL   = "https://tw-envisoft.tedp.vn"
API_BASE   = "https://admin-qttd.tedp.vn"
BASIC_USER = "tw-admin"
BASIC_PASS = "tw-admin"
FORM_USER  = "duongngocbach"
FORM_PASS  = "1234567890!@#$%^&*()"

# Data API endpoint (discovered via network interception)
DATA_ENDPOINT = f"{API_BASE}/api/eos/data-average-by-time/exceed-by-time"

# Stations list endpoint
STATIONS_ENDPOINT = (
    f"{API_BASE}/api/stations/search/"
    "findByStationTypeAndProvinceIdAndFtpConnectionStatusAndStatus"
    "AndUsingStatusAndStationNameAndTenantCode"
)

PAGE_SIZE = 200  # rows per page for data fetch


# ─── Step 1: Login via Playwright, keep context alive for API calls ────────────
def login_and_get_context(playwright):
    """
    Opens browser, bypasses Layer 1 (HTTP Basic Auth) and Layer 2 (web form),
    navigates to admin-qttd.tedp.vn to establish the WAF session, then returns
    (browser, context, api_request_context).

    The WAF (F5 BIG-IP) ties authentication to the TLS fingerprint + x-bni-fpc
    cookie pair. Python requests uses a different TLS stack and always gets 403.
    Using context.request keeps everything in Chromium's network stack.
    """
    print("[AUTH] Starting browser login ...")

    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context(
        http_credentials={"username": BASIC_USER, "password": BASIC_PASS}
    )
    page = context.new_page()

    # ── Layer 1: embed credentials in URL to bypass HTTP Basic Auth ────────────
    print("[AUTH] Navigating to login page ...")
    page.goto(
        f"https://{BASIC_USER}:{BASIC_PASS}@tw-envisoft.tedp.vn/",
        wait_until="domcontentloaded",
        timeout=60000,
    )
    page.wait_for_timeout(1500)

    # ── Layer 2: web form login ────────────────────────────────────────────────
    print("[AUTH] Submitting Layer 2 credentials ...")
    try:
        page.wait_for_selector(".blockUI", state="hidden", timeout=15000)
    except Exception:
        pass
    page.get_by_role("textbox", name="Tên người dùng:").fill(FORM_USER)
    page.get_by_role("textbox", name="Mật khẩu:").fill(FORM_PASS)
    page.evaluate("document.querySelector('input[type=submit]').click()")
    page.wait_for_load_state("domcontentloaded")

    # ── Wait for SPA to load and redirect to admin-qttd.tedp.vn ───────────────
    print("[AUTH] Waiting for redirect to admin-qttd.tedp.vn ...")
    page.wait_for_timeout(10000)

    # ── Navigate to the data page — this establishes the WAF session ───────────
    print("[AUTH] Establishing WAF session via data-average-by-time page ...")
    page.goto(
        f"{MAIN_URL}/eos/view_log/average_data_by_time",
        wait_until="domcontentloaded",
        timeout=30000,
    )
    page.wait_for_timeout(5000)

    # Log cookies for diagnosis
    cookies = context.cookies()
    admin_cookies = [c for c in cookies if "admin-qttd" in c.get("domain", "")]
    print(f"[AUTH] Session established. admin-qttd.tedp.vn cookies: "
          f"{[c['name'] for c in admin_cookies]}")

    # Keep the page open and reuse it for all API calls via browser navigation.
    # context.request / Python requests both use a different TLS stack from
    # Chromium and get 403 from the F5 BIG-IP WAF. Browser page navigation
    # uses the correct TLS fingerprint and the established WAF session.
    return browser, context, page


# ─── API helper: GET via browser navigation (bypasses WAF TLS check) ──────────
def api_get(page: Page, url: str, params: dict = None) -> dict:
    """
    Navigate the browser to an API URL and parse the JSON body.
    This uses Chromium's TLS fingerprint + the established WAF session,
    which is the only combination the F5 BIG-IP WAF accepts.

    We use response.text() (raw HTTP body) rather than document.body.innerText
    to avoid Chrome's JSON viewer interfering with the text extraction.
    """
    full_url = url + ("?" + urlencode(params) if params else "")
    response = page.goto(full_url, wait_until="domcontentloaded", timeout=60000)
    if response is None:
        raise RuntimeError(f"Navigation returned no response: {full_url}")
    if response.status not in (200, 404):
        raise RuntimeError(f"HTTP {response.status}: {full_url}")
    if response.status == 404:
        raise RuntimeError("404")
    text = response.text()
    return json.loads(text)


# ─── Step 2: Fetch station list ────────────────────────────────────────────────
def get_all_stations(page: Page,
                     station_type: str = "KK",
                     province_id: str = "") -> list:
    """
    Fetches all stations filtered by station_type (e.g. 'KK') and province_id.
    Paginates automatically.
    """
    all_stations = []
    pg = 0
    print(f"[STATIONS] Fetching stations (type={station_type or 'all'}) ...")

    while True:
        params = {
            "stationType":  station_type,
            "stationName":  "",
            "areaIds":      province_id,
            "page":         pg,
            "size":         50,
        }
        try:
            data = api_get(page, STATIONS_ENDPOINT, params)
        except Exception as e:
            print(f"[ERROR] Stations API failed: {e}")
            break

        stations = data.get("_embedded", {}).get("stations", [])
        if not stations:
            break

        all_stations.extend(stations)
        page_info   = data.get("page", {})
        total_pages = page_info.get("totalPages", 1)
        print(f"  Page {pg+1}/{total_pages}: {len(stations)} stations")
        if pg >= total_pages - 1:
            break
        pg += 1

    print(f"[STATIONS] Total found: {len(all_stations)}")
    return all_stations


# ─── Step 3: Fetch average data for one station ────────────────────────────────
def fetch_station_data(page: Page,
                       station: dict,
                       from_date: str,
                       to_date: str,
                       view_type: str = "minute",
                       data_type: int = 1) -> list:
    """
    Calls the data API for a single station and paginates through all results.

    Parameters
    ----------
    view_type : str
        'minute' | 'hour' | '8hour' | 'day' | 'month'
    data_type : int
        1 = original data
    """
    station_id   = station.get("id", "")
    station_name = station.get("stationName") or station.get("name") or station_id

    all_rows = []
    pg       = 0

    # API expects datetime in ISO format
    from_dt = f"{from_date}T00:00:00"
    to_dt   = f"{to_date}T23:59:59"

    while True:
        params = {
            "stationId": station_id,
            "fromDate":  from_dt,
            "toDate":    to_dt,
            "viewType":  view_type,
            "dataType":  data_type,
            "page":      pg,
            "size":      PAGE_SIZE,
        }
        try:
            data = api_get(page, DATA_ENDPOINT, params)
        except json.JSONDecodeError:
            # Non-JSON response (e.g. 404 HTML page) means no data
            break
        except Exception as e:
            print(f"    [WARN] Request failed for {station_name}: {e}")
            break

        # Handle paginated response shapes
        if isinstance(data, list):
            rows        = data
            total_pages = 1
        elif "content" in data:
            rows        = data["content"]
            total_pages = data.get("totalPages", 1)
        elif "_embedded" in data:
            key         = list(data["_embedded"].keys())[0]
            rows        = data["_embedded"][key]
            total_pages = data.get("page", {}).get("totalPages", 1)
        else:
            rows        = [data] if data else []
            total_pages = 1

        # Annotate each row with station metadata
        for row in rows:
            row["station_name"] = station_name
            row["station_id"]   = station_id
            row["station_code"] = station.get("stationCode") or station.get("code") or ""

        all_rows.extend(rows)

        if pg >= total_pages - 1:
            break
        pg += 1

    return all_rows


# ─── Step 4: Save rows to Excel ────────────────────────────────────────────────
def save_to_excel(rows: list, filepath: Path, sheet_name: str = "Data"):
    if not rows:
        print(f"  [SKIP] No data — {filepath.name}")
        return 0

    df = pd.DataFrame(rows)

    # Move station metadata columns to front
    meta = [c for c in ["station_name", "station_code", "station_id"] if c in df.columns]
    rest = [c for c in df.columns if c not in meta]
    df   = df[meta + rest]

    df.rename(columns={
        "station_name": "Station Name",
        "station_code": "Station Code",
        "station_id":   "Station ID",
    }, inplace=True)

    sheet_name = sheet_name[:31]  # Excel sheet name limit

    with pd.ExcelWriter(str(filepath), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        ws = writer.sheets[sheet_name]

        # Style header
        header_fill = PatternFill("solid", fgColor="1F6AA5")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    print(f"  [SAVED] {filepath.name}  ({len(df)} rows)")
    return len(df)


# ─── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Fetch AQMS average data from tw-envisoft.tedp.vn"
    )
    parser.add_argument("--from-date", required=True,
                        help="Start date YYYY-MM-DD  (e.g. 2026-03-03)")
    parser.add_argument("--to-date", required=True,
                        help="End date YYYY-MM-DD    (e.g. 2026-03-04)")
    parser.add_argument("--output-dir", default="./output",
                        help="Directory for Excel output (default: ./output)")
    parser.add_argument("--station-type", default="KK",
                        help="Station type filter: KK=Air, NT=Wastewater, NM=Surface water (default: KK)")
    parser.add_argument("--province-id", default="",
                        help="Filter by province ID (default: all provinces)")
    parser.add_argument("--view-type", default="minute",
                        choices=["minute", "hour", "8hour", "day", "month"],
                        help="Data averaging period (default: minute)")
    parser.add_argument("--combined", action="store_true",
                        help="Save all stations in one Excel file (one sheet per station)")
    args = parser.parse_args()

    # Validate dates
    try:
        datetime.strptime(args.from_date, "%Y-%m-%d")
        datetime.strptime(args.to_date,   "%Y-%m-%d")
    except ValueError:
        print("[ERROR] Dates must be in YYYY-MM-DD format")
        sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as playwright:
        # ── 1. Login — keep browser alive for all API calls ───────────────────
        browser, context, page = login_and_get_context(playwright)

        try:
            # ── 2. Stations ───────────────────────────────────────────────────
            stations = get_all_stations(page,
                                        station_type=args.station_type,
                                        province_id=args.province_id)
            if not stations:
                print("[ERROR] No stations found. Check filters or authentication.")
                sys.exit(1)

            # ── 3. Fetch data ─────────────────────────────────────────────────
            print(f"\n[DATA] Date range : {args.from_date}  →  {args.to_date}")
            print(f"[DATA] View type  : {args.view_type}")
            print(f"[DATA] Stations   : {len(stations)}\n")

            if args.combined:
                # ── All stations → one workbook, one sheet per station ────────
                out_file = output_dir / f"AQMS_{args.station_type}_{args.from_date}_{args.to_date}.xlsx"
                with pd.ExcelWriter(str(out_file), engine="openpyxl") as writer:
                    for i, station in enumerate(stations, 1):
                        name = (station.get("stationName") or
                                station.get("name") or
                                station.get("id", f"station_{i}"))
                        print(f"[{i:>3}/{len(stations)}] {name}")

                        rows = fetch_station_data(page, station,
                                                  args.from_date, args.to_date,
                                                  view_type=args.view_type)
                        if not rows:
                            print(f"         → No data")
                            continue

                        df         = pd.DataFrame(rows)
                        meta       = [c for c in ["station_name","station_code","station_id"] if c in df.columns]
                        df         = df[meta + [c for c in df.columns if c not in meta]]
                        sheet_name = re.sub(r"[^\w\s\-]", "", name)[:31]

                        df.to_excel(writer, index=False, sheet_name=sheet_name)
                        print(f"         → {len(df)} rows  →  sheet '{sheet_name}'")

                print(f"\n[DONE] Combined file: {out_file}")

            else:
                # ── One Excel file per station ────────────────────────────────
                total_rows = 0
                for i, station in enumerate(stations, 1):
                    name = (station.get("stationName") or
                            station.get("name") or
                            station.get("id", f"station_{i}"))
                    print(f"[{i:>3}/{len(stations)}] {name}")

                    rows = fetch_station_data(page, station,
                                              args.from_date, args.to_date,
                                              view_type=args.view_type)

                    safe = re.sub(r'[<>:"/\\|?*]', "_", name).strip()[:60]
                    out  = output_dir / f"{safe}_{args.from_date}_{args.to_date}.xlsx"
                    total_rows += save_to_excel(rows, out, sheet_name="Average Data")

                print(f"\n[DONE] {len(stations)} files saved to '{output_dir}'  |  Total rows: {total_rows}")

        finally:
            browser.close()


if __name__ == "__main__":
    main()
